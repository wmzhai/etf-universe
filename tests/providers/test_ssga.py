from io import BytesIO

from openpyxl import Workbook

from etf_universe.providers.ssga import parse_ssga_workbook


def test_parse_ssga_workbook_extracts_holdings_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Fund Name:", "SPDR S&P 500 ETF Trust"])
    sheet.append(["Ticker Symbol:", "SPY"])
    sheet.append([])
    sheet.append([])
    sheet.append(["As of Mar 28, 2026", None])
    sheet.append([])
    sheet.append(["Ticker", "Name", "Weight"])
    sheet.append(["AAPL", "Apple Inc.", "6.1"])
    sheet.append(["BRK/B", "Berkshire Hathaway Inc.", "1.9"])

    buffer = BytesIO()
    workbook.save(buffer)

    result = parse_ssga_workbook(buffer.getvalue(), "https://example.com/spy.xlsx")

    assert result.source_format == "xlsx"
    assert result.as_of_date.isoformat() == "2026-03-28"
    assert [row.constituent_symbol for row in result.rows] == ["AAPL", "BRK/B"]
    assert result.profile is not None
    assert result.profile.fundName == "SPDR S&P 500 ETF Trust"


def test_parse_ssga_workbook_skips_footer_and_tickerless_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append([])
    sheet.append(["As of Mar 28, 2026", None])
    sheet.append([])
    sheet.append(["Ticker", "Name", "Weight"])
    sheet.append(["AAPL", "Apple Inc.", "6.1"])
    sheet.append([None, "US DOLLAR", "0.05"])
    sheet.append(["MSFT", "Microsoft Corp.", None])
    sheet.append([None, "The holdings are subject to change.", None])
    sheet.append(["BRK/B", "Berkshire Hathaway Inc.", "1.9"])

    buffer = BytesIO()
    workbook.save(buffer)

    result = parse_ssga_workbook(buffer.getvalue(), "https://example.com/spy.xlsx")

    assert [row.constituent_symbol for row in result.rows] == ["AAPL", "BRK/B"]
    assert [row.weight for row in result.rows] == [6.1, 1.9]


def test_parse_ssga_workbook_discovers_date_and_header_without_fixed_indexes() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SPDR ETF Holdings"])
    sheet.append([])
    sheet.append(["Generated at", "2026-03-31T10:00:00"])
    sheet.append(["Holdings:", "As of 30-Mar-2026"])
    sheet.append(["Preamble row"])
    sheet.append(["Name", "Ticker", "Identifier", "Weight", "Sector"])
    sheet.append(["Apple Inc.", "AAPL", "037833100", "6.1", "Information Technology"])
    sheet.append(["Berkshire Hathaway Inc.", "BRK/B", "084670702", "1.9", "Financials"])

    buffer = BytesIO()
    workbook.save(buffer)

    result = parse_ssga_workbook(buffer.getvalue(), "https://example.com/spy.xlsx")

    assert result.as_of_date.isoformat() == "2026-03-30"
    assert [row.constituent_symbol for row in result.rows] == ["AAPL", "BRK/B"]
