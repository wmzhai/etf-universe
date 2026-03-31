from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from etf_universe.contracts import EtfSpec, FetchResult, SourceHoldingRow
from etf_universe.normalization import clean_text, parse_date_from_text
from etf_universe.providers.base import HTTP_TIMEOUT, build_source_row, get_by_header


def _find_as_of_date(rows: list[tuple[Any, ...]]):
    for row in rows:
        for cell in row:
            cell_text = clean_text(cell)
            if cell_text is None or "as of" not in cell_text.casefold():
                continue
            try:
                return parse_date_from_text(r"As of\s+(.+)", cell_text)
            except ValueError:
                continue

    raise ValueError("Unable to find as-of date in SSGA workbook")


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(rows):
        headers = [clean_text(value) for value in row]
        index = {header: i for i, header in enumerate(headers) if header}
        if {"Ticker", "Name", "Weight"}.issubset(index):
            return row_idx, index

    raise ValueError("Unable to find required headers in SSGA workbook")


def parse_ssga_workbook(content: bytes, source_url: str) -> FetchResult:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    as_of_date = _find_as_of_date(rows)
    header_row_idx, index = _find_header_row(rows)

    records: list[SourceHoldingRow] = []

    for row in rows[header_row_idx + 1 :]:
        if not any(row):
            continue

        name = get_by_header(row, index, "Name")
        symbol = get_by_header(row, index, "Ticker")
        if clean_text(name) is None and clean_text(symbol) is None:
            continue

        parsed = build_source_row(
            constituent_symbol=symbol,
            constituent_name=name,
            weight=get_by_header(row, index, "Weight"),
            asset_class="Equity",
            security_type="Common Stock",
        )
        if parsed.constituent_symbol is None or parsed.weight is None:
            continue

        records.append(parsed)

    return FetchResult(
        as_of_date=as_of_date,
        source_url=source_url,
        source_format="xlsx",
        rows=records,
    )


def fetch_ssga(spec: EtfSpec, session) -> FetchResult:  # noqa: ANN001
    response = session.get(spec.source_url, timeout=HTTP_TIMEOUT)
    response.raise_for_status()
    return parse_ssga_workbook(response.content, spec.source_url)
